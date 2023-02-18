import os
import openpyxl


path = os.chdir("D:/Education/xls_files/")
list_of_xl_files = os.listdir(path)
# print("Список всех файлов: ", list_of_xl_files)


def info_about_count():     # Информация о кол-ве листов в каждом найденном файле xl
    for n in list_of_xl_files:
        tbl = openpyxl.open(n)
        sheet = tbl.sheetnames
        sheet_active = tbl[sheet[0]]
        print(f'\nВ файл "{n}" входят следующие листы:')
        count = 0
        for i in tbl:
            count += 1
            print(f'{count}.{i}')


def files_in_folder():
    list_of_files = []

    for file_names in list_of_xl_files:
        file = f"D:/Education/xls_files/{file_names}"
        list_of_files.append(file)
    return list_of_files


# def find_entering_value(list_of_files):
#
#     for tab in list_of_files:
#         complete_table = []
#         open_file = openpyxl.open(filename=tab, read_only=True)
#         sheet = open_file.active
#
#         for row in sheet.iter_rows():
#             for column in row:
#                 complete_table.append(str(column.value))
#         if enter in complete_table:
#             print(f"\033[0;42m OK \033[0;0m Значение есть в файле по адресу: \033[1;33m{tab}\033[0;0m")
#         else:
#             print(f"\033[0;41m NO \033[0;0m Значения нет в файле по адресу: \033[1;33m{tab}\033[0;0m")

def find_entering_value(list_of_files):

    for tab in list_of_files:
        complete_table = []
        open_file = openpyxl.open(filename=tab)
        sheet = open_file.sheetnames

        for i in range(len(sheet)):
            count = ''
            sheet_active = open_file[sheet[i]]
            count += str(i)
            for row in sheet_active.iter_rows():
                for column in row:
                    complete_table.append(str(column.value))
            if enter in complete_table:
                print(f'\033[0;42m OK \033[0;0m Значение есть в файле по адресу: \033[1;33m{tab}\033[0;0m, в листе: '
                      f'\033[1;33m"{sheet[int(count)]}"\033[0;0m')
            else:
                print(f"\033[0;41m NO \033[0;0m Значения нет в файле по адресу: \033[1;33m{tab}\033[0;0m")

# def enter_value():
#     global enter
#     enter = input("Введите искомое значение: ")
#     print(f'\nПоиск значения: \033[1;33m"{enter}"\033[0;0m...\n')
#     try:
#         enter = float(enter)
#     except (ValueError, TypeError):
#         pass # enter = str(enter)


enter = input("Введите искомое значение: ")
find_entering_value(files_in_folder())

